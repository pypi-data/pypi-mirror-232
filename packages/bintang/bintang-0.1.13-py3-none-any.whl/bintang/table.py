#from bintang.core import Bintang
from openpyxl import load_workbook
from bintang.column import Column
from bintang.cell import Cell
from bintang.row import Row
import json
import sqlite3
import uuid
import re
import types
import sys
import unicodedata
from thefuzz import process as thefuzzprocess
from bintang.log import log
# import logging

# log = logging.getLogger(__name__)
# FORMAT = "[%(filename)s:%(lineno)s - %(funcName)10s() ] %(message)s"
# logging.basicConfig(format=FORMAT)
# log.setLevel(logging.DEBUG)

INDEX_COLUMN_NAME = 'idx'
PARENT_PREFIX = ''

class ColumnNotFoundError(Exception):
    def __init__(self,table, column):
        self.message = "Cannot find column '{}' in table {}.".format(column, table)
        super().__init__(self.message)
    
        

class Table(object):
    """Define a Bintang table object
       - provide columns to store a dictionary of column objects
       - provide rows to store a dictionary of row objects
    """
    def __init__(self,name, bing=None):
        self.bing = bing
        self.name = name
        self.__columns = {}
        self.__rows = {}
        self.__temprows = []
        self.__last_assigned_columnid= -1 #
        self.__last_assigned_rowid = -1 # for use when row created
        self.__last_assigned_idx = -1 # for use when add idx
        self.__be = None
        


    def __getitem__(self, idx): # subscriptable version of self.get_row_asdict()
        #return self.__rows[idx] # bad design. a raw row isn't that useful at client code
        return self.get_row_asdict(idx)


    def __setitem__(self, rowcell, value): # subscriptable version of self.update_row()
        # where arg rowcell is a tuple of an idx of rows & column  passed by client code
        # client code signature: table_obj[idx,column] = value
        self.update_row(rowcell[0], rowcell[1], value)
        

    def __repr__(self):
        tbl = {}
        tbl['name'] = self.name
        columns = []
        for k,v in self.__columns.items():
            columns.append(dict(id=v.id, name=v.name))
        tbl['columns'] = columns
        return json.dumps(tbl, indent=2)


    def __len__(self):
        """ return the length of rows"""    
        return len(self.__rows)
            

    def set_to_sql_colmap(self, columns):
        if isinstance(columns, list):
            return dict(zip(columns, columns))
        elif isinstance(columns, dict):
            return columns
        

    def to_sql(self, conn, schema, table, columns, method='prep', max_rows = 1):
        if method == 'prep':
            return self.to_sql_prep(conn, schema, table, columns, max_rows=max_rows)
        elif method =='string':
            return self.to_sql_string(conn, schema, table, columns, max_rows=max_rows)

 
    def to_sql_string(self, conn, schema, table, columns, max_rows = 300):
        colmap = self.set_to_sql_colmap(columns)
        src_cols = [x for x in colmap.values()]
        dest_columns = [x for x in colmap.keys()]
       

        sql_template = 'INSERT INTO {}.{} ({}) VALUES'
        str_stmt = sql_template.format(schema,table,",".join(['"{}"'.format(x) for x in colmap]))

        sql_cols_withtype = self.set_sql_datatype(dest_columns, conn, schema, table)
        sql_cols_withliteral = self.set_sql_literal(sql_cols_withtype, conn)
        
        # start insert to sql
        cursor = conn.cursor()
        temp_rows = []  
        total_rowcount = 0 # to hold total record affected
        for idx, values in self.iterrows(src_cols, row_type='list'):
            ## question: can values and d_cols_withliteral align? dict python 3.7 is ordered and will solve it?
            sql_record = self.gen_sql_literal_record(values, sql_cols_withliteral)
            temp_rows.append(sql_record)
            if len(temp_rows) == max_rows:
                stmt = str_stmt + ' {}'.format(",".join(temp_rows))
                #log.debug(stmt)
                cursor.execute(stmt)
                total_rowcount += cursor.rowcount
                temp_rows.clear()
        if len(temp_rows) > 0:
            stmt = str_stmt + ' {}'.format(",".join(temp_rows))
            #log.debug(stmt)
            cursor.execute(stmt)
            total_rowcount += cursor.rowcount
        return total_rowcount           


    def gen_sql_literal_record(self, values, sql_cols_withliteral):
        sql_record = []
        sql_columns = [x for x in sql_cols_withliteral.keys()]
        for i, value in enumerate(values):
            col = sql_columns[i]
            literals = sql_cols_withliteral[col]
            sql_value = self.gen_sql_literal_value(literals, value)
            sql_record.append(sql_value)
        return "({})".format(','.join(sql_record))
        

    def gen_sql_literal_value(self, literals, value):
        if value == "" or value is None:
            return "NULL"
        if isinstance(value, str):
            value = value.replace("'","''")
        return "{}{}{}".format('' if literals[0] is None else literals[0], value, '' if literals[1] is None else literals[1])
    

    def get_sql_typeinfo_table(self, conn):
        cursor = conn.cursor()
        sql_type_info_tuple = cursor.getTypeInfo(sqlType = None)
        columns_ = [column[0] for column in cursor.description]
        tobj = Table('sql_typeinfo')
        for row in sql_type_info_tuple:
            tobj.insert(row, columns_)        
        return tobj
    

    def set_sql_datatype(self, dest_columns, conn, schema, table):
        cursor = conn.cursor()
        sql_columns = cursor.columns(schema=schema, table=table)
        columns_ = [column[0] for column in cursor.description]
        tobj = Table('sql_columns_')
        for row in sql_columns: #cursor.columns(schema=schema, table=table):
            tobj.insert(row, columns_)
        sql_columns_withtype = {}
        for col in dest_columns:
            _type = tobj.get_value('type_name', where = lambda row: row['column_name']==col)
            sql_columns_withtype[col] = _type
        return sql_columns_withtype
    

    def set_sql_literal(self, sql_cols_withtype, conn):
        sql_typeinfo_tab = self.get_sql_typeinfo_table(conn)
        sql_cols_withliteral = {}
        for k, v in sql_cols_withtype.items():
            prefix = sql_typeinfo_tab.get_value('literal_prefix',where=lambda row: row['type_name']==v)
            suffix = sql_typeinfo_tab.get_value('literal_suffix',where=lambda row: row['type_name']==v)
            literals = (prefix,suffix)
            sql_cols_withliteral[k] = literals
        return sql_cols_withliteral


    def to_sql_prep(self, conn, schema, table, columns, max_rows = 1):
        if max_rows <= len(self): # validate max_row
            mrpb = max_rows # assign max row per batch
        else:
            #log.warning('Warning! max_rows {} set greater than totalrows {}. max_rows set to 1'.format(max_rows, len(self)))
            mrpb = 1
        numof_col = len(columns) # num of columns

        colmap = self.set_to_sql_colmap(columns)
        src_cols = [x for x in colmap.values()]
        dest_columns = [x for x in colmap.keys()]
        
        # create as prepared statement
        sql_template = 'INSERT INTO {}.{} ({}) VALUES {}'
        param_markers = self.gen_row_param_markers(numof_col, mrpb)
        prep_stmt = sql_template.format(schema,table,",".join(['"{}"'.format(x) for x in dest_columns]),param_markers)
        cursor = conn.cursor()
        temp_rows = []
        total_rowcount = 0
        for idx, row in self.iterrows(src_cols, row_type='list'):
            for v in row:
                temp_rows.append(v)
            if len(temp_rows) == (mrpb * numof_col):
                log.debug(prep_stmt)
                log.debug(temp_rows)
                cursor.execute(prep_stmt, temp_rows)
                total_rowcount += cursor.rowcount
                temp_rows.clear()     
        
        if len(temp_rows) > 0: # if any reminder
            param_markers = self.gen_row_param_markers(numof_col, int(len(temp_rows)/numof_col))
            prep_stmt = sql_template.format(schema,table,",".join(['"{}"'.format(x) for x in dest_columns]),param_markers)
            cursor.execute(prep_stmt, temp_rows)
            total_rowcount += cursor.rowcount
        return total_rowcount        


    def gen_row_param_markers(self,numof_col,num_row):
        param = "(" + ",".join("?"*numof_col) + ")"
        params = []
        for i in range(num_row):
            params.append(param)
        return ",".join(params)
    


    def gen_create_sqltable(self, dbms):
        # scanning table to get column properties
        # and assign the data type and size (when able)
        self.set_data_props()
        for col in self.get_columns():
            cobj = self.__columns[self.get_columnid(col)]
            if cobj.data_type is None:
                if len(cobj.data_props) == 0: # a column that has no data at all
                    cobj.data_type = 'str'  # force to str
                    cobj.column_size = 50 # force to 50
                # loop through the data props for column that has it
                if 'str' in cobj.data_props:
                    cobj.data_type = 'str'
                    col_size = cobj.data_props['str']['column_size']
                    # loop through any other type if the column_size bigger, if it is assign it.
                    for k, v in cobj.data_props.items():
                        if v['column_size'] > col_size:
                            col_size = v['column_size']
                    cobj.column_size = col_size #cobj.data_props['str']['column_size']
                elif 'datetime' in cobj.data_props:
                    cobj.data_type = 'datetime'
                    cobj.column_size = cobj.dataprops['datetime']['column_size']
                else: # just get the first one and break. to be observed later on!
                    for type, prop in cobj.data_props.items():
                        cobj.data_type = type
                        break
        
        # use type_map to translate the type
        create_columns = []
        for col in self.get_columns():
            cobj = self.__columns[self.get_columnid(col)]
            if cobj.data_type == 'str':
                create_item = ['[{}]'.format(cobj.name), type_map[dbms][cobj.data_type],'({})'.format(cobj.column_size)]
                create_columns.append(create_item)
            else:
                create_item = ['[{}]'.format(cobj.name), type_map[dbms][cobj.data_type]]
                create_columns.append(create_item)
        create_columns_str = []
        for i, citem in enumerate(create_columns):
            create_item_str = []
            if i == 0: # add tab if first, the rest will be added at later join
                create_item_str.append('\t')
            for i in citem:
                create_item_str.append(i)
            create_item_str.append('\n')
            create_columns_str.append(' '.join(create_item_str))
        create_sqltable_templ = "CREATE TABLE {} (\n{})".format(self.name, '\t,'.join(create_columns_str))
        return create_sqltable_templ 
    

    def add_column(self, name, data_type=None, column_size=None):
        # check if the passed name already exists
        columnid = self.get_columnid(name)
        if columnid is None:
            cobj = Column(name)
            if data_type is not None:
                cobj.data_type = data_type
            if column_size is not None:
                cobj.column_size = column_size
            cobj.id = self.__last_assigned_columnid + 1
            self.__columns[cobj.id] = cobj
            self.__last_assigned_columnid= self.__last_assigned_columnid + 1


    def update_column(self,name, data_type=None, column_size=None):
        # check if the passed name already exists
        columnid = self.get_columnid(name)
        if columnid is not None:
            if data_type is not None:
                self.__columns[columnid].data_type = data_type
            if column_size is not None:
                self.__columns[columnid].column_size = column_size


    def _add_column(self,name):
        # check if the passed name already exists
        columnid = self._get_columnid(name)
        log.debug(columnid)
        if columnid is None:
            cobj = Column(name)
            cobj.id = self.__last_assigned_columnid + 1
            self.__columns[cobj.id] = cobj
            self.__last_assigned_columnid= self.__last_assigned_columnid + 1        


    def get_columnid(self,column):
        for id, cobj in self.__columns.items():
            # match the column case insensitive
            if cobj.get_name_uppercased() == column.upper():
                return id
        return None
        

    
    def get_columnids(self,columns=None):
        columnids = []
        if columns is None: # assume user want all available column ids
            return [id for id in self.__columns.keys()]
        for column in columns:
            columnid = self.get_columnid(column)
            if columnid is None:
                raise ValueError('Cannot find column name {}.'.format(column))
            else:
                columnids.append(columnid)
        return columnids

            
    def rename_column(self,old_column,new_column):
        for v in self.__columns.values():
            if v.name == old_column:
                v.name = new_column
                return

        

    def drop_column(self,column):
        # get columnid
        columnid = self.get_columnid(column)
        #provide warning if the passed column does not exist
        if columnid is None:
            log.warning("warning... trying to drop a non-existence column '{}'".format(column))
            return False
        # delete the cell from cell
        for row in self.__rows.values():
            row.cells.pop(columnid,None)
        # delete the column
        self.__columns.pop(columnid,None)


    def get_column(self,columnid):
        return self.__columns[columnid].name
        

    def get_columns(self, columns=None):
        return [x.name for x in self.__columns.values()]


    def _get_columnnames_lced(self, columns=None):
        return {x.name.lower(): x.name  for x in self.__columns.values()}


    def get_data_props(self, column):
        columnid = self.get_columnid(column)
        return self.__columns[columnid].data_props


    def get_column_object(self, column):
        columnid = self.get_columnid(column)
        return self.__columns[columnid]

 
    def validate_column(self, column):
        """return column as the one stored in table.columns"""
        if column.lower() in self._get_columnnames_lced().keys():
            return self._get_columnnames_lced().get(column.lower())
        else:
            raise ColumnNotFoundError(self.name, column)


    def validate_columns(self, columns):
        """return columns from those stored in table.columns"""
        res = []
        for column in columns:
            if column.lower() in self._get_columnnames_lced().keys():
                res.append(self._get_columnnames_lced().get(column.lower()))
            else:
                raise ColumnNotFoundError(self.name, column)
        return res


    def check_column(self, column):
        """check if column exits in table.columns"""
        if column.lower() not in self._get_columnnames_lced().keys():
            extracted = thefuzzprocess.extract(column, self.get_columns(), limit=2)
            fuzzies = [x[0] for x in extracted if x[1]>85]
            raise ValueError ('could not find column {}. Did you mean {}?'.format(column,' or '.join(fuzzies)))
            


    def VOID_validate_columns(self,columns):
        res = []
        for column in columns:
            columnid = self.get_columnid(column)
            if columnid is not None:
                existing_column = self.get_column(columnid)
                res.append(existing_column)
            else:
                raise ValueError("Cannot find column {}.".format(column))
        return columns


    def get_index(self, condition_column, condition_value):
        """will return the first index which value matches the condition"""
        for idx, row in self.iterrows():
            # test if string
            # if type(row[condition_columnname])==str:pass
            if row[condition_column] == condition_value:
                return idx

    def get_indexes(self, condition_column, condition_value):
        """will return a list of indexes which value matches the condition"""
        indexes = []
        for idx, row in self.iterrows():
            # test if string
            # if type(row[condition_columnname])==str:pass
            if row[condition_column] == condition_value:
                indexes.append(idx)
        return indexes    


    def get_value_OLD(self, search_column, condition_column, condition_value):
        ## will return a scalar value for the first match
        for idx, row in self.iterrows():
            if row[condition_column] == condition_value:
                # print(row[condition_columnname])
                # print(row[search_columnname])
                # print('hello')
                return row[search_column]


    def get_value_original(self, column, where):
        """ 
        return a scalar value.
        switch different process if user want column pass as string or lambda
        params:
        column: either column name or using lambda expression
        where: always a lambda expression"""
        
        if isinstance(column,str):
            for idx, row in self.iterrows():
                if where(row):
                    return row[column]
        elif isinstance(column, types.FunctionType):
            for idx, row in self.iterrows():
                if where(row):
                    return column(row)
        else:
            raise 'param column must be either column name or lambda and param where must be a lambda.'


    def get_value(self, column, where = None):
        """ 
        return a scalar value.
        switch different process if user want column pass as string or lambda
        params:
        column: either column name or using lambda expression
        where: always a lambda expression"""
        if where is not None:
            if isinstance(column,str):
                for idx, row in self.iterrows():
                    if where(row):
                        return row[column]
            elif isinstance(column, types.FunctionType):
                for idx, row in self.iterrows():
                    if where(row):
                        return column(row)
            else:
                raise 'param column must be either column name or lambda and param where must be a lambda.'
        else:
            if isinstance(column,str):
                for idx, row in self.iterrows():
                    return row[column]
            elif isinstance(column, types.FunctionType):
                for idx, row in self.iterrows():
                    return column(row)
            else:
                raise 'param column must be either column name or lambda and param where must be a lambda.'


                
        



    def make_row(self,id=None, option=None):
        """make a new row.
        by default it sets uuid4 for id.
        """
        if id is None and option is None:
            row = Row(self.__last_assigned_rowid + 1)
            self.__last_assigned_rowid += 1 #increment rowid
        elif id is not None and option is None:
            row = Row(id)
        elif id is None and option == 'uuid':
            row = Row(uuid.uuid4())
        return row


    def insert(self, record, columns=None):
        """ restrict arguments for record insertion for this function as the followings:
        1. a pair of columns and its single-row values
        deprecated2. a pair of columns and its multiple-row values (as a list of tuple)
        """

        if isinstance(record, dict):
            row = self.make_row()
            for idx, (col, val) in enumerate(record.items()):
                cell = self.make_cell(col,val)
                row.add_cell(cell) # add to row
            self.add_row(row)                                    
        elif isinstance(columns,list) or isinstance(columns,tuple) or isinstance(record,list) or isinstance(record,tuple):
            row = self.make_row()
            for idx, col in enumerate(columns):
                cell = self.make_cell(col,record[idx])
                row.add_cell(cell) # add to rows
            if self.__be is not None:
                self.__temprows.append(json.dumps({v.columnid: v.value for v in row.cells.values()}))
                if len(self.__temprows) == self.__be.max_row_for_sql_insert:
                    self.add_row_into_be()
            else:
                self.add_row(row)    
        else:
            raise ValueError("Arg for record set for dictionary or list/tuple of values with list/tuple of columns.")
            
            
    def insert_old(self,columns,values):
        """ restrict arguments for data insertion for this function as the followings:
        1. a pair of columns and its single-row values
        deprecated2. a pair of columns and its multiple-row values (as a list of tuple)
        """

        # if isinstance(values[0],tuple): # if no. 2
        #     for value in values:
        #         if isinstance(value,tuple): # if True then expect a multi-row insert. so add a row for each value
        #             row = self.make_row()
        #             for idx,column in enumerate(columns):
        #                 cell = self.make_cell(column,value[idx])
        #                 row.add_cell(cell)
        #             # add to rows
        #             self.add_row(row)    
        #else: # if no. 1return None
        # elif isinstance(columns,list) and not isinstance(values[0],tuple):
        if isinstance(columns,list) or isinstance(columns,tuple) or isinstance(values,list) or isinstance(values,tuple):
            row = self.make_row()
            for idx, column in enumerate(columns):
                cell = self.make_cell(column,values[idx])
                row.add_cell(cell)
            # add to rows
            if self.__be is not None:
                self.__temprows.append(json.dumps({v.columnid: v.value for v in row.cells.values()}))
                if len(self.__temprows) == self.__be.max_row_for_sql_insert:
                    self.add_row_into_be()
            else:
                self.add_row(row)    
        else:
            raise ValueError("insert only allows a pair of columns and values with type list or tuple")


    def add_row_into_be(self):
        self.__be.add_row(self.name, self.__temprows)
        self.__temprows.clear()      


    def insert_dict(self, adict):
        self.insert([x for x in adict.keys()], [x for x in adict.values()])


    def insert_todb(self,columns,values):
        """ restrict arguments for data insertion for this function as the followings:
        1. a pair of columns and its single-row values
        2. a pair of columns and its multiple-row values (as a list of tuple)
        """
        if isinstance(columns,list) and not isinstance(values[0],tuple):
            row = self.make_row()
            for idx, column in enumerate(columns):
                cell = self.make_cell(column,values[idx])
                row.add_cell(cell)
            # add to rows
            self.add_row(row)
        else:
            raise ValueError("Insert only allows a pair of columns and values in a list or tuple")

    def get_indexes_OLD(self):
        return [x for x in self.__rows.keys()]


    def get_rowidx_byrowid(self, rowid):
        debug = False
        # if debug:
        #     print("\n  ------------------in get_row_rowid (table.py) --------------------")
        for idx, row in self.__rows.items():
            if debug:
                print(idx, row)
            if row.id == rowid:
                return idx
        # if debug:
        #     print("\n  ------------------out get_row_rowid (table.py) -------------------")
        return None    


    def upsert_table_path_row(self, tprow):
        #log.debug("\n  ------------------in upsert_table_path_row (table.py) --------------------")
        # extract the rowid and use it as the table index (the key of rows{})
        # create a row if the rowid not found in the table's index
        res_idx = self.get_rowidx_byrowid(tprow.id)
        
        if res_idx is None:
            #log.debug(f'inserting... row does not exist {tprow}')
            row = self.make_row(tprow.id)
            # re-make cells from tprow
            for id, c in tprow.cells.items():
                #log.debug(f'{id} cell: {c}')
                if c.is_key == True:
                    cell = self.make_cell(PARENT_PREFIX + c.get_column(), c.value)
                else:
                    cell = self.make_cell(c.get_column(), c.value)

                row.add_cell(cell)
            # add to rows
            self.add_row(row)
            
        elif res_idx is not None:
            #log.debug(f"updating... row exists {tprow}")
            for id, c in tprow.cells.items():
                if c.is_key == True:
                    cell = self.make_cell(PARENT_PREFIX + c.get_column(), c.value)
                else:
                    cell = self.make_cell(c.get_column(), c.value)
                self.__rows[res_idx].add_cell(cell)        
        #log.debug("\n  ------------------out upsert_table_path_row (table.py)-------------------")


    def make_cell(self,column,value,new_column=True):
        columnid = self.get_columnid(column)
        if columnid is None: # if columnid is None then assume user wants a new column
            if new_column == True:
                self.add_column(column)
                columnid = self.get_columnid(column) # reassign the columnid
                if self.__be is not None:
                    self.__be.add_column(self.name, columnid, column)
                # deprecated moved up columnid = self.get_columnid(column) # reassign the columnid
        if columnid is None:
            raise ValueError("Cannot make cell due to None column name.")    
        return Cell(columnid,value)


    def add_row_deprecated(self,row):
        rows_idx = len(self.__rows) # can cause re-assign a deleted row
        self.__rows[rows_idx] = row

    def add_row(self, row):
        rows_idx = self.__last_assigned_idx + 1
        self.__last_assigned_idx += 1
        self.__rows[rows_idx] = row


    def _XXgen_row_asdict(self, row, columns, rowid=False):
        res = {}
        if rowid == True:
            res['_rowid_'] = row.id # add rowid for internal purpose eg. a merged table
        for column in columns:
            columnid = self.get_columnid(column)
            if columnid not in row.cells:
                res[column] = None
            else:
                res[column] = row.cells[columnid].value
        return res   
    

    def delete_row(self,index):
        self.__rows.pop(index,None)


    def delete_rows(self,indexes):
        for index in indexes:
            self.delete_row(index)


    def get_row(self,idx):
        if idx in self.__rows:
            return self.__rows[idx]
        else:
            raise KeyError ('Cannot find index {}.'.format(idx))


    def get_valid_columnname(self, column):
        columnid = self.get_columnid(column)   # column in this line passed by user
        column = self.get_column(columnid) # ensure the same column passed as result.
        return column        
    

    def get_row_asdict(self, idx, columns=None, rowid=False):
        if idx not in self.__rows:
            raise KeyError ('Cannot find index {}.'.format(idx))
        if idx in self.__rows:
            if columns is None:
                columns = self.get_columns()
            if columns is not None:
                columns = self.validate_columns(columns)
            return self._gen_row_asdict(self.__rows[idx],columns, rowid)
            

    def _gen_row_asdict(self, row, columns, rowid=False):
        res = {}
        if rowid == True:
            res['rowid_'] = row.id # add rowid for internal purpose eg. a merged table
        for column in columns:
            # get valid column. these two lines defined in def get_valid_columnname()
            columnid = self.get_columnid(column)   # column in this line passed by user
            column = self.get_column(columnid) # ensure the same column passed as result.
            if columnid not in row.cells:
                res[column] = None
            else:
                res[column] = row.cells[columnid].value        
        return res


    def get_row_aslist(self, idx, columns=None):
        if columns is None:
            columns = self.get_columns()
        columnids = self.get_columnids(columns)
        return self._gen_row_aslist(self.__rows[idx],columnids)


    def _gen_row_aslist(self, row, columnids):
        return row.get_values(columnids)

       
    def iterrows(self, 
                 columns: list=None, 
                 row_type: str='dict', 
                 where=None, 
                 rowid: bool=False):
        # validate user's args
        if columns is not None:
            for column in columns:
                self.check_column(column)
        if columns is None:
                columns = self.get_columns() # assign all available column names
        if row_type == 'dict': 
            if self.__be is None:
                if where is not None:
                    for idx, row in self.__rows.items():
                        if where(self._gen_row_asdict(row, columns, rowid)):
                            yield idx, self._gen_row_asdict(row,columns,rowid)
                else:
                    for idx, row in self.__rows.items():
                        yield idx, self._gen_row_asdict(row,columns,rowid)
            if self.__be is not None:
                for idx, row in self.__be.iterrows_asdict(self.name, columns):
                    yield idx, row
        elif row_type == 'list':
            columnids = self.get_columnids(columns)
            for idx, row in self.__rows.items():
                yield idx, self._gen_row_aslist(row,columnids)


    def set_data_props(self):
        """ scan table to obtain columns properties - data type, column size (if str type then the max of len of string)"""
        columnids = self.get_columnids()
        for idx, row in self.__rows.items():
            for columnid in columnids:
                # log.debug(row)
                # log.debug('id: {}, value: {}'.format(idx , self.__columns[columnid].name))
                if columnid in row.cells:
                    self.set_data_props_datatype(columnid, row.cells[columnid].value)
                    if row.cells[columnid].value is not None:
                        if type(row.cells[columnid].value).__name__ == 'str':
                            self.set_data_props_str_column_size(columnid, row.cells[columnid].value)


    def set_data_props_datatype(self, columnid,value):
        # set a column's data_type list
        if value is not None:
            if type(value).__name__ not in self.__columns[columnid].data_props:
                self.__columns[columnid].data_props[type(value).__name__] = dict(column_size=1)
                if type(value).__name__ == 'int':
                    self.__columns[columnid].data_props['int']['column_size'] = 10
                if type(value).__name__ == 'float':
                    self.__columns[columnid].data_props['float']['column_size'] = 15
                if type(value).__name__ == 'bool':
                    self.__columns[columnid].data_props['bool']['column_size'] = 1
                if type(value).__name__ == 'datetime':
                    self.__columns[columnid].data_props['datetime']['column_size'] = 19        

                
    def set_data_props_str_column_size(self, columnid, value):
        if len(value) > self.__columns[columnid].data_props['str']['column_size']:
            self.__columns[columnid].data_props['str']['column_size'] = len(value)       
    
        
    def print(self,top=None, columns=None):
        rows_values = []
        for enum,row in enumerate(self.__rows.values(),1):
            row_values = []
            for columnid in self.__columns.keys():
                cell_value = None
                if columnid in row.cells:
                    cell_value = row.cells[columnid].value
                row_values.append(cell_value)
            rows_values.append(row_values)
            if enum == top:
                break
        # get and print out columns
        columns = [x.name for x in self.__columns.values()]
        print(columns) 
        # get and print out rowresult_as
        for row in rows_values:
            print(row)


    def print_aslist(self, top=None, columns=None):
        print("idx",self.get_columns(columns))
        for idx, row in self.iterrows(row_type='list',columns=columns):
            print(idx, row)


    def print_asdict(self, top=None, columns=None):
        for idx, row in self.iterrows(columns=columns):
            print(idx, row)

    
    def columnnames_valid(self, columns):
        existing_columnnames = self._get_columnnames_lced()
        for column in columns:
            if column.lower() not in existing_columnnames:
                raise ValueError("Cannot find column name {}.".format(column))
        return True


    def replace_insensitively(self, old, repl, text):
        return re.sub('(?i)'+re.escape(old), lambda m: repl, text)

    def validate_stmt(self, stmt):
        invalid_keywords = ['import ']
        for ik in invalid_keywords:
            if ik in stmt:
                raise ValueError("Found invalid keyword {} in the statement!".format(repr(ik)))


    def set_cellvalue_stmt(self, row, stmt):
        columns = re.findall('`(.*?)`',stmt) # extract column names from within a small tilde pair ``
        for column in columns:
            stmt = stmt.replace('`' + column + '`',repr(row[column]))
        return stmt


    def exec_stmt(self, stmt):
        # log.debug('-' * 10)
        # log.debug(stmt)
        res = {'retval':False}
        # res["__retval"] = None
        try:
            exec(stmt, globals(), res)
        except Exception as e:
            res["error"] = e
        finally:
            return res


    def get_index_exec(self, stmt):
        """return the fist index that matches the condition."""
        columnname_case_insensitive = False
        if columnname_case_insensitive == True:
            columnnames_in_stmt = re.findall('`(.*?)`',stmt) # extract column names from within a small tilde pair ``
            self.columnnames_valid(columnnames_in_stmt)  # validate column name from the stmt
            columnnames_in_stmt = [self.get_valid_columnname(x) for x in columnnames_in_stmt]
            for column in columnnames_in_stmt:
                valid_columnname = self.get_valid_columnname(column)
                stmt = self.replace_insensitively(column, valid_columnname, stmt)
        # scan the rows to search the index
        indexes = []
        for idx, row in self.iterrows():
            stmt_set = self.set_cellvalue_stmt(row, stmt)
            self.validate_stmt(stmt_set)
            ret = self.exec_stmt(stmt_set)
            if ret['retval'] == True:
                return idx
        

    def get_indexes_exec(self, stmt):
        """return a list of indexes that match the condition."""
        columnname_case_insensitive = False
        if columnname_case_insensitive == True:
            columnnames_in_stmt = re.findall('`(.*?)`',stmt) # extract column names from within a small tilde pair ``
            self.columnnames_valid(columnnames_in_stmt)  # validate column name from the stmt
            columnnames_in_stmt = [self.get_valid_columnname(x) for x in columnnames_in_stmt]
            for column in columnnames_in_stmt:
                valid_columnname = self.get_valid_columnname(column)
                stmt = self.replace_insensitively(column, valid_columnname, stmt)
        # scan the rows to search the index
        indexes = []
        for idx, row in self.iterrows():
            stmt_set = self.set_cellvalue_stmt(row, stmt)
            self.validate_stmt(stmt_set)
            ret = self.exec_stmt(stmt_set)
            if ret['retval'] == True:
                indexes.append(idx)
        return indexes


    def delete_rows_by_stmt(self, stmt):
        indexes = self.get_index_exec(stmt)
        self.delete_rows(indexes)


    def update_row(self,idx,column,value):
        self.__rows[idx].add_cell(self.make_cell(column,value))


    def update_row_all(self,column,value):
        
        for idx in self.__rows:
            self.__rows[idx].add_cell(self.make_cell(column,value))    


    def exists(self, where):
        for idx, row in self.iterrows():
            try:
                if where(row):
                    return True
            except:
                print(row)
                pass
        return False

    def exists_original(self, where):
        for idx, row in self.iterrows():
            try:
                if where(row):
                    return True
            except:
                print(row)
                pass
        return False    
        
        
    def update(self, column, value, where=None):
        if isinstance(value, types.FunctionType):
            # value passed as function
            for idx, row in self.iterrows():
                if where is None:
                    try:
                        val = value(row) # function called
                        self.update_row(idx, column, val)
                    except Exception as e:
                        log.warning(e)
                        pass    
                else:    
                    try:
                        if where(row):
                            val = value(row) # function called
                            self.update_row(idx, column, val) 
                    except Exception as e:
                        log.warning(e)
                        pass 
        else:
            # value passed as non function eg. an int or str
            for idx, row in self.iterrows():
                if where is None:
                    self.update_row(idx, column, value)
                else:    
                    try:
                        if where(row): # function called 
                            self.update_row(idx, column, value) 
                    except Exception as e:
                        # print(e)
                        pass 



    
    def get_index_lambda(self,expr):
        for idx, row in self.iterrows():
            log.debug('idx {} row{}:'.format(idx, row))
            ret = expr(row)
            if ret:
                return idx


    def get_indexes_lambda(self,expr):
        indexes = []
        for idx, row in self.iterrows():
            log.debug('idx {} row{}:'.format(idx, row))
            ret = expr(row)
            if ret:
                indexes.append(idx)       
        return indexes


    def filter(self, expr, columns=None):
        tobj = Table('mybing3')
        if columns is None:
            for idx, row in self.iterrows():
                try:
                    if expr(row):
                        columns = tuple(row.keys())
                        values = tuple(row.values())
                        tobj.insert(columns, values)
                except:
                    pass        
        else:
            columns_ = self.validate_columns(columns)
            for idx, row in self.iterrows():
                try:
                    if expr(row):
                        values = self.get_row_aslist(idx,columns_)
                        tobj.insert(columns_,values)
                except:
                    pass        
        return tobj


    def innerjoin(self
                ,rtable # either string or Table object
                ,on: list[tuple] # list of lkey & r key tuple
                ,into: str
                ,out_lcolumns: list=None
                ,out_rcolumns: list=None
                ,rowid=False):
        
        rtable_obj = None
        if isinstance(rtable,Table):
            rtable_obj = rtable
            rtable = rtable.name
        else:
            rtable_obj = self.bing[rtable]
        
        # validate input eg. column etc
        lkeys = [x[0] for x in on] # generate lkeys from on (1st sequence)
        rkeys = [x[1] for x in on] # generate rkeys from on (2nd sequence)
        lkeys = self.validate_columns(lkeys)
        #rkeys = self[rtable].validate_columns(rkeys)
        rkeys = rtable_obj.validate_columns(rkeys)
        if out_lcolumns is not None:
            out_lcolumns = self.validate_columns(out_lcolumns)
        if out_rcolumns is not None:
            # out_rcolumns = self[rtable].validate_columns(out_rcolumns)
            out_rcolumns = rtable_obj.validate_columns(out_rcolumns)

        # resolve columns conflicts
        rcol_resolved = self._resolve_join_columns(rtable_obj, rowid)
        # create an output table
        out_table = into
        out_tobj = self.bing.create_table(out_table)
        out_tobj = self.bing.get_table(out_table)
        
        # for debuging create merged table to store the matching rowids
        #merged = self.create_table("merged",["lrowid","rrowid"])

        numof_keys = len(on) #(lkeys)
        # loop left table
        for lidx, lrow in self.iterrows(columns=lkeys, rowid=rowid):
            # loop right table
            for ridx, rrow in rtable_obj.iterrows(columns=rkeys, rowid=rowid):
                matches = 0 # store matches for each rrow
                # compare value for any matching keys, if TRUE then increment matches
                for i in range(numof_keys): 
                    if _match_caseless_unicode(lrow[lkeys[i]], rrow[rkeys[i]]):
                        matches += 1 # incremented!
                if matches == numof_keys: # if fully matched, create the row & add into the output table
                    #debug merged.insert(["lrowid","rrowid"], [lrow["_rowid"], rrow["_rowid"]])
                    #and add the row to output table
                    outrow = out_tobj.make_row()
                    # add cells from left table
                    log.debug(f'{lidx}')
                    outrow = self._add_lcell(lidx,  outrow, out_table, out_lcolumns, rowid)
                    # add cells from right table
                    outrow = self._add_rcell(ridx, rtable_obj, outrow, out_table, out_rcolumns, rcol_resolved, rowid)   
                    out_tobj.add_row(outrow)
        #debug merged.print() 
        return out_tobj


    def _resolve_join_columns (self, rtable_obj, rowid=False):
        # resolve any conflict column by prefixing its tablename
        # notes: conflict column occurs when the same column being used in two joining table.
        rcol_resolved = {}
        if rowid:
            rcol_resolved['rowid_'] = rtable_obj.name + '_' + 'rowid_'
        lcolumns = self.get_columns()
        for col in rtable_obj.get_columns():
            if col in lcolumns:
                rcol_resolved[col] = rtable_obj.name + '_' + col
            else:
                rcol_resolved[col] = col
        return rcol_resolved
    
    def _add_lcell(self, lidx, outrow, out_table, out_lcolumns, rowid):
        for k, v in self.get_row_asdict(lidx,rowid=rowid).items():
            if out_lcolumns is None:
                # incude all columns
                cell = self.bing[out_table].make_cell(k,v)
                outrow.add_cell(cell)
            if out_lcolumns is not None:
                # include only the passed columns
                if k in out_lcolumns:
                    cell = self.bing[out_table].make_cell(k,v)
                    outrow.add_cell(cell)
        return outrow            


    def _add_rcell(self, ridx, rtable, outrow, out_table, out_rcolumns, rcol_resolved, rowid):
        for k, v in rtable.get_row_asdict(ridx,rowid=rowid).items():
            if out_rcolumns is None:
                cell = self.bing[out_table].make_cell(rcol_resolved[k],v)
                outrow.add_cell(cell)
            if out_rcolumns is not None:
                # include only the passed columns
                if k in out_rcolumns:
                    cell = self.bing[out_table].make_cell(rcol_resolved[k],v)
                    outrow.add_cell(cell)
        return outrow
    # def get_type_used(self, column):
    #     # iterate through the table and suss out the type
    #     used_types = []
    #     for idx, row in self.iterrows(columns=[column]):
    #         #if row[column] is not None:
    #         used_type = type(row[column]).__name__
    #         if used_type not in used_types:
    #             used_types.append(used_type)
    #     return used_types


    # def get_types_used(self, column):
    #     # self.__columns[1].name = 'manuf'
    #     # for k,v in self.__columns.items():
    #     #     log.debug(k,v.name)
        
    #     # quit()
    #     # iterate through the table and suss out the type

    #     for idx, row in self.iterrows():
    #         #if row[column] is not None:
    #         used_type = type(row[column]).__name__
    #         if used_type not in used_types:
    #             used_types.append(used_type)
    #     return used_types    


    def blookup(self, 
                lkp_table: object, 
                on: str, 
                ret_columns: list[str] | list[tuple]):
        lkp_table_obj = None
        if isinstance(lkp_table,Table):
            lkp_table_obj = lkp_table
            lkp_table = lkp_table.name
        else:
            lkp_table_obj = self.bing[lkp_table]
        
        # validate input eg. column etc
        lkeys = [x[0] for x in on] # generate lkeys from on (1st sequence)
        rkeys = [x[1] for x in on] # generate rkeys from on (2nd sequence)
        lkeys = self.validate_columns(lkeys)
        rkeys = lkp_table_obj.validate_columns(rkeys)
        lcolumn_prematch = self.get_columns() # will use this list when matching occurs below
        # validate ret_columns once:
        # otherwise will make lots of call later
        valid_ret_columns = []
        for item in ret_columns:
            if isinstance(item,str):
                item = lkp_table_obj.validate_column(item)
                valid_ret_columns.append(item)
            if isinstance(item, tuple):
                valid_column  = lkp_table_obj.validate_column(item[0])
                item_ = tuple([valid_column,item[1]])
                valid_ret_columns.append(item_)
        numof_keys = len(on)
        for lidx, lrow in self.iterrows(lkeys, rowid=True):
            for ridx, rrow in lkp_table_obj.iterrows():
                matches = 0
                for i in range(numof_keys):
                    # if lrow[lkeys[i]] == rrow[rkeys[i]]:
                    if _match_caseless_unicode(lrow[lkeys[i]], rrow[rkeys[i]]):
                        matches += 1 # increment
                if matches == numof_keys:
                    # update this table lrow for each ret_columns
                    for item in valid_ret_columns:
                        if isinstance(item, str): # if item is a column
                            value = lkp_table_obj[ridx][item]
                            if item in lcolumn_prematch:
                                print('item',item,'in',self.name)
                                self.update_row(lidx, lkp_table_obj.name + '_' + item, value)
                            else:
                                self.update_row(lidx, item, value)
                        if isinstance(item, tuple): # if a tuple (0=column to return from lkp_table, 1=as_column)
                            value = lkp_table_obj[ridx][item[0]]
                            self.update_row(lidx, item[1], value)                   


    def groupbycount(self, column):
        res_dict = {}
        for idx, row in self.iterrows(column):
            value = next(iter(row.values()))
            if value not in res_dict:
                res_dict[value] = 1
            else:
                res_dict[value] += 1
        return res_dict
    
    
    
    def groupby2count(self, columns):
        res_dict = {} #key=a tuple of columns, value=count
        for idx, row in self.iterrows(columns, row_type='list'):
            # DEPRECATED - No need tobe a string trow = tuple([str(x or 'None') for x in row])
            trow = tuple(row)
            keys = res_dict.keys()
            if trow not in res_dict.keys():
                res_dict[trow] = 1
            else:
                res_dict[trow] += 1       
        return res_dict
    
    
    def DEV_groupby3count(self, columns):
        group_tobj = Table('test')
        res_dict = {} #key=a tuple of columns, value=count
        for idx, row in self.iterrows(columns, row_type='list'):
            trow = tuple(row)
            if trow not in res_dict.keys():
                res_dict[trow] = 1
            else:
                res_dict[trow] += 1
                
        for k, v in res_dict.items():
            columns_ = columns + ['count']
            values = list(k) + [v]
            group_tobj.insert(columns_, values)       
        return group_tobj
    

    def DEV_groupby_count(self, columns, count_column):
        group_tobj = Table('testq')
        res_dict = {} #key=a tuple of columns, value=count
        for idx, row in self.iterrows(columns, row_type='list'):
            trow = tuple(row)
            if trow not in res_dict.keys():
                res_dict[trow] = 1
            else:
                res_dict[trow] += 1      
        for k, v in res_dict.items():
            columns_ = columns + [count_column]
            values = list(k) + [v]
            group_tobj.insert(values, columns_,)       
        return group_tobj
    

    def DEV_groupby(self, columns, save_as):
        res_dict = {} #key=a tuple of columns, value=count
        for idx, row in self.iterrows(columns, row_type='list'):
            trow = tuple(row)
            if trow not in res_dict.keys():
                res_dict[trow] = 1
            else:
                res_dict[trow] += 1
        # create a table object
        tobj = self.bing.create_table(save_as)
        tobj = self.bing.get_table(save_as)
        # loop dict and insert to table object
        for k, v in res_dict.items():
            print(k,v)
            columns_toinsert = columns + ['count']
            print('columns_toinsert',columns_toinsert)
            values = list(k) + [v]
            print('values',values)
            tobj.insert(values, columns_toinsert,)
        return tobj
    

    def read_excel(self, path, sheetname):
        wb = load_workbook(path, read_only=True, data_only=True)
        ws = wb[sheetname]
        columns = []
        Nonecolumn_cnt = 0
        for rownum, row_cells in enumerate(ws.iter_rows(),start=1):
            values = [] # hold column value for each row
            if rownum == 1:
                for cell in row_cells:
                    if cell.value is None:
                        columname = 'noname' + str(Nonecolumn_cnt)
                        Nonecolumn_cnt += 1
                        columns.append(columname)
                    else:
                        columns.append(cell.value)
                if Nonecolumn_cnt > 0:
                    log.warning('Warning! Noname column detected!')          
            
            if rownum > 1:
                for cell in row_cells:
                    values.append(cell.value)
                # if rownum == 370:
                #     log.debug(f'{values} at rownum 370.')
                #     log.debug(any(values))
                if any(values):
                    self.insert(values, columns)
        if self.__be is not None:
            self.add_row_into_be()


    def read_sql(self, conn, sql_str=None, params=None ):
        cursor = conn.cursor()
        if sql_str is None:
            sql_str = "SELECT * FROM {}".format(self.name)
        if params is not None:
            cursor.execute(sql_str, params)
        else:
            cursor.execute(sql_str)
        columns = [col[0] for col in cursor.description]
        for row in cursor.fetchall():
            self.insert(row, columns)        


    def reindex(self):
        # dict is already indexed with python 3.7+
        # this only means we rebuild row id so any deleted idx can be reused
        idxs = list(self.__rows)
        for i, idx in enumerate(idxs):
            self.__rows[i] = self.__rows[idx] # reassign
            if i != idx:
                del self.__rows[idx]


    def to_csv(self, path, columns=None, index=False, delimiter=',',\
               quotechar='"', quoting=3): #csv.QUOTE_NONE is 3
        import csv
        if columns is None:
            columns = self.get_columns()
        with open(path, 'w', newline = '\n') as csvfile:
            csvwriter = csv.writer(csvfile, dialect='excel', delimiter=delimiter,
                                   quotechar=quotechar, quoting=quoting)
            columns_towrite = [col for col in columns]
            if index:                       # if column index wanted
                idx_col = INDEX_COLUMN_NAME
                if isinstance(index, str):  # if user wanted own index column name
                    idx_col = index
                columns_towrite.insert(0,idx_col)
            # write header to csvfile
            csvwriter.writerow(columns_towrite)
            if index:
                for idx, row in self.iterrows(columns, row_type='list'):
                    csvwriter.writerow([idx] + row)
            else:
                for idx, row in self.iterrows(columns, row_type='list'):
                    csvwriter.writerow(row)
            

    def to_dict(self, columns=None):
        if columns is None:
            columns = self.get_columns()
        res = {}
        res['table'] = self.name
        res['columns'] = columns
        res['rows'] = []
        # add row
        for idx, row in self.iterrows(columns):
            res['rows'].append(row)
        return res    
            

    def to_excel(self, path, columns=None, index=False):
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        # add header
        columns_towrite = []
        if columns is None:
            columns_towrite = self.get_columns()
        else:
            columns_towrite = [col for col in columns]
        log.debug('index: {}'.format(index))
        if index:
            if index:                          # if column index wanted
                idx_column = INDEX_COLUMN_NAME
                if isinstance(index, str):     # if user wanted own index column name
                    idx_column = index      
                columns_towrite.insert(0,idx_column)
            if index != True:
                columns_towrite.insert(0,str(index))        
        ws.append(columns_towrite)
        # add row
        if index:
            for idx, row in self.iterrows(columns, row_type='list'):
                row.insert(0,idx)
                ws.append(row)
        if not index:
            for idx, row in self.iterrows(columns, row_type='list'):
                ws.append(row)          
        wb.save(path)


    def to_list(self, column):
        if isinstance(column, str):
            res_list = []
            for idx, row in self.iterrows([column]):
                res_list.append(next(iter(row.values())))
            return res_list    
        else:
            raise ValueError('column must be a single column string!')    


    def _to_json_file(self, path, columns=None):
        """ this will be useful for big table
        and we are going to write row by row into a file
        instead a one-oof write"""
        if columns is None:
            columns = self.get_columns()
        with open(path, 'w') as f:
            f.write('{') # write opening json obj
            f.write('{}'.format(json.dumps(self.name)))
            f.write(':')
            f.write('[') # write opening array

            for idx, row in self.iterrows(columns=columns):
                print(json.dumps(row))
                f.write(json.dumps(row))
                if not (idx + 1) == len(self): 
                    f.write(',') # if not eof then write obj seperator
            f.write(']') # write closing array
            f.write('}') # write closing json obj


    def to_json(self):
        """This is just a placeholder.
        Converting dict to Json is trivia by using Json dumps/dump.
        No need a function to write here"""
        

class Table_Path(Table):
    def __init__(self, name):
        super().__init__(name)
        self.path = name
        self.children = []        


    def __repr__(self):
        tbl = {}
        tbl['name'] = self.name
        tbl['path'] = self.path
        columns = []
        for k,v in self._Table__columns.items():
            columns.append(dict(id=v.id, name=v.name))
        tbl['columns'] = columns
        return json.dumps(tbl, indent=2)


    def get_path_aslist(self):
        path_as_list = []
        if self.path == '/': # if root
            path_as_list.append('/')
            return path_as_list
        for node in self.path.split("/"):
            if node == '': # if the 1st one. 1st node have discarded by split
                path_as_list.append('/')
            else:
                path_as_list.append(node)
        return path_as_list        

def _match_primitive(value1, value2):
    "Ascii case sensitive matching"
    if value1 == value2:
        return True
    
def _match_caseless(value1, value2):
    "Ascii ase insensitive matching"
    if isinstance(value1, str) and isinstance(value2, str):
        if value1.lower() == value2.lower():
            return True
    elif isinstance(value1, str) or isinstance(value2, str):
        if str(value1) == str(value2):
            return True    
    else:
        if value1 == value2:
            return True    
        
def _match_caseless_unicode(value1, value2):
    "Unicode case insensetive matching"
    if isinstance(value1, str) and isinstance(value2, str):
        if _normalize_caseless(value1) == _normalize_caseless(value2):
            return True
    elif isinstance(value1, str) or isinstance(value2, str):
        if str(value1) == str(value2):
            return True
    else:
        if value1 == value2:
            return True

def _normalize_caseless(string):
    return unicodedata.normalize("NFKD", string.casefold())

type_map = {
    'sqlserver': {
                'str':'nvarchar'
                ,'int':'int'
                ,'datetime':'datetime'
                ,'float':'float'
                ,'bool':'bit'
                ,'bytes':'varbinary'
                }
    }


