# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell  # openpyxl == 3.0.7,需要去__init__文件中import cell

from excel_tools.read import XlsxReader

import os
from typing import Optional, Union, Generator, List, Tuple, Any

from excel_tools.read.base_reader import ExcelBaseObject


class XlsxWrite(XlsxReader):
    def __init__(self, name: str, path: Optional[str] = ''):
        super(XlsxWrite, self).__init__(name=name, path=path)

    def open_xl(self):
        path = os.path.join(self.path, self.name)
        try:
            open(path, 'r')
        except FileNotFoundError:
            self._xl = openpyxl.Workbook()
        else:
            self._xl = openpyxl.load_workbook(path)

    def write_value_to_cell(self, value: Union[int, str, bool], rowx: int, colx: int, style=None):
        """
        将value内容，填充到指定单元格

        Args:
            value: 需要填入的值
            rowx: 对应行数
            colx: 对应列数
            style: 指定单元格样式

        Returns:
            None
        """
        cell = self.get_cell(rowx=rowx, colx=colx)
        cell.value = value

        # TODO: style
        if style and cell.style != style.name:
            cell.style = style

    def write_value_to_col(self, value: Union[list, tuple], colx: int, start_rowx: Optional[int] = 1, style=None):
        """
        将value内容，填充到指定列

        Args:
            value: 需要填入表格中的数据
            colx: 指定列数
            start_rowx: 左切片行数
            style: 指定单元格样式

        Returns:
            None
        """
        if not isinstance(value, (list, tuple)):
            raise ValueError('write_value_to_col error, 传入错误的值', value)
        for v in value:
            self.write_value_to_cell(value=v, colx=colx, rowx=start_rowx, style=style)
            start_rowx += 1

    def write_value_to_row(self, value: Union[list, tuple], rowx: int, start_colx: Optional[int] = 1, style=None):
        """
        将value内容，填充到指定行

        Args:
            value: 需要填入表格中的数据
            rowx: 行数
            start_colx: 左切片列数
            style: 指定单元格样式

        Returns:
            None
        """
        if not isinstance(value, (list, tuple)):
            raise ValueError('write_value_to_col error, 传入错误的值', value)
        for v in value:
            self.write_value_to_cell(value=v, colx=start_colx, rowx=rowx, style=style)
            start_colx += 1

    def save(self):
        """ 保存工作簿"""
        self._xl.save(os.path.join(self.path, self.name))

    def if_condition(self) -> bool:
        """
        判断当前excel是否可以读写

        Returns:
            bool
        """
        if os.access(os.path.join(self.path, self.name), os.X_OK):
            try:
                self.save()
                closed = True
            except PermissionError:
                closed = False
            return closed