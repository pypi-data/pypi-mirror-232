# -*- coding: utf-8 -*-
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.cell import Cell  # openpyxl == 3.0.7,需要去__init__文件中import cell

import os
from typing import Optional, Union, Generator, List, Tuple, Any

from excel_tools.read.base_reader import ExcelBaseObject


class XlsxReader(ExcelBaseObject):
    def __init__(self, sheet: Union[int, str, None] = 0, *args, **kwargs):
        """
        xlsx读取类，可以读取以下格式的表格：xlsx

        Args:
            name(str): 文件名
            path(str): 文件路径

            sheet(int,str): 需要读取的工作表(表名或表索引)
        """
        super(XlsxReader, self).__init__(*args, **kwargs)
        self.set_sheet(sheet=sheet)

    def open_xl(self):
        self._xl = openpyxl.load_workbook(os.path.join(self._path, self._name))

    def set_sheet(self, sheet: Union[int, str, None] = 0) -> None:
        if isinstance(sheet, str):
            self._sheet = self._xl[sheet]
        elif isinstance(sheet, int):
            self._sheet = self._xl[self._xl.sheetnames[sheet]]
        else:
            self._sheet = self._xl.active

    @property
    def sheet_name(self) -> str:
        return self._sheet.title

    @property
    def max_row(self) -> int:
        if not hasattr(self, '_max_row'):
            setattr(self, '_max_row',  self._sheet.max_row)

        return getattr(self, '_max_row')

    @max_row.setter
    def max_row(self, rowx: int) -> None:
        setattr(self, '_max_row', rowx)

    @property
    def max_column(self) -> int:
        if not hasattr(self, '_max_col'):
            setattr(self, '_max_col',  self._sheet.max_column)

        return getattr(self, '_max_col')

    @max_column.setter
    def max_column(self, colx: int) -> None:
        setattr(self, '_max_col', colx)

    def get_rows(self) -> Generator[Tuple[Cell], Any, None]:
        return self._sheet.rows

    def get_row(self, rowx: int, start_colx: Optional[int] = 0, end_colx: int = None) -> Tuple[Cell]:
        end_colx = end_colx if end_colx is not None and end_colx <= self.max_column else self.max_column
        return tuple(self._sheet.iter_rows(min_row=rowx, max_row=rowx, min_col=start_colx, max_col=end_colx))[0]

    def get_row_len(self, rowx: int, start_colx: Optional[int] = 0, end_colx: int = None) -> int:
        return len(self.get_row(rowx=rowx, start_colx=start_colx, end_colx=end_colx))

    def get_row_value(self, rowx: int, start_colx: Optional[int] = 0, end_colx: int = None) -> List[Any]:
        return [cell.value for cell in self.get_row(rowx=rowx, start_colx=start_colx, end_colx=end_colx)]

    def get_col(self, colx: int, start_rowx: Optional[int] = 0, end_rowx: Optional[int] = None) -> Tuple[Cell]:
        if end_rowx:
            if end_rowx > self.max_row:
                end_rowx = self.max_row
        else:
            end_rowx = self.max_row

        return self._sheet[get_column_letter(colx)][start_rowx:end_rowx]

    def get_col_value(self, colx: int, start_rowx: Optional[int] = 0, end_rowx: Optional[int] = None) -> List[Any]:
        return [cell.value for cell in self.get_col(colx=colx, end_rowx=end_rowx, start_rowx=start_rowx)]

    def get_cell(self, rowx: int, colx: int) -> Cell:
        return self._sheet.cell(rowx, colx)

    def get_cell_value(self, rowx: int, colx: int) -> Any:
        return self.get_cell(rowx, colx).value
